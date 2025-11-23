# 
# LINHA DE EXECUÇÃO a partir de G:\Meu Drive\Cursos e Treinamentos\Cientista de Dados\Treinamento Python\Consulta receita despesa\consulta_receita_despesa\Scripts
# 
# TEMPO DE EXECUÇÃO: 5 MINUTOS
#
# EXISTE NO MÉTODO MAIN, UMA MANEIRA DE SE FORÇAR CARGAS.
#
# OS MESES VÃO SENDO INSERIDOS NO BI, A MEDIDA QUE VÃO SENDO ENCERRADOS
#
# 3 TABELAS SERÃO ESCRITAS: TBL_COB_FENASEG e TBL_TMP_RECEITA_SNG e TBL_TMP_FLUXO_CAIXA_SNG
#
# NOVA FORMA DE EXECUTAR: Dentro da pasta C:\Users\qlik_admin\Scripts\consulta_receita_despesa\Scripts como administrador 
# python.exe consulta_receita_sng_siseg.py
# 
# SOMENTE PARA ESSAS CONTAS DO FLUXO DE CAIXA: 030103 - > SISEG, 030101 -> SNG
#
# Consultas por mês e ano de pagamento. Cuidado, diferente de competencia, que é o ulitmo dia, do mês anterior ao pagamento
#
#
from pandas import DataFrame
from pandas import to_datetime
from pandas import read_excel
from pandas import concat
from requests import post
from datetime import timedelta, date, datetime
from sys import argv
from time import time
import oracledb
import openpyxl
from urllib3 import disable_warnings
from openpyxl import load_workbook
import pysftp as sftp
import os
import re
import sqlalchemy as sa
import warnings

warnings.filterwarnings('ignore','.*Failed to load.*')

# PARA RETIRAR A MENSAGEM DE WARNING DE VERIFICAÇÃO DE CERTIFICADO
disable_warnings()

tempo_inicial = (time())  # em segundos

class ParamErro(Exception):
    pass

class AnoErro(Exception):
    pass

class MesErro(Exception):
    pass

class FlxErro(Exception):
    pass

class EmpErro(Exception):
    pass

class ErroAPI(Exception):
    pass

class ErroGeralAPI(Exception):
    pass

class NotitErro(Exception):
    pass

class Mesinvalido(Exception):
    pass

class Anoinvalido(Exception):
    pass

#class Competencia(Exception):
#    pass

def formatar_valor(valor):    
    
    if valor == '':
        
        #print('Número do Título com desconto e valor '': ',numerodotitulo)
                
        valor = float(str(valor).replace('','0'))
        valor = f'{valor:,.2f}'.replace('.', ',')
        
    elif valor == '0':
        valor = float(valor)
        valor = str(f'{valor:,.2f}').replace('.', ',')
            
    else:
        valor = float(str(valor).replace(',','.'))
        valor = f'{valor:_.2f}'.replace('.', ',').replace('_', '.') 
        
        #print('Número do Título com desconto: ',numerodotitulo)  
        #print('Valor diferente de nada: ',valor)       
    
    #print('Valor: ',valor) 
               
    return valor  


def formatar_quantidade(quantidade):    
    
    if quantidade == '':        
        quantidade = int(str(quantidade).replace('','0'))       
    
    else:
        
        quantidade = int(quantidade)
                
        quantidade = f'{quantidade:_}'.replace('_','.')
        #print ('Quantidade: ',quantidade)        
            
    return quantidade


# OBTENDO TODOS OS CENTROS DE CUSTO
def obter_centros_custo(indice,interfacecontaspagarreceber):
    
    cursor = conexao.cursor()
    codigoccusto = []
        
    for l in indice:
        
        for i in interfacecontaspagarreceber[l]['InterfaceGrupoPagarReceber']:
        
            codigo = i['NumerodoCentrodeCusto']
            codigoccusto.append(codigo)
    
    codigoccusto = tuple(set(codigoccusto))
    
    print('Tamanho: ',len(codigoccusto))    

    if len(codigoccusto) == 1:
        codigoccusto = f"('{codigoccusto[0]}')"
    
    print('Código ccusto: ', codigoccusto)
    
    # OBTENDO TODOS OS CENTROS DE CUSTO
    sql = f"SELECT distinct cc_codigo,cc_descricao FROM MXM_PRD.ccusto_cc WHERE cc_ativo = 'S' and cc_codigo IN {codigoccusto}"
    
    cccusto_cc_cursor = cursor.execute(sql)

    listacentrosdecusto = cccusto_cc_cursor.fetchall() # LISTA DE TUPLAS 

    #print('Lista centros de custo')
    #print(listacentrosdecusto)
    
    for r in listacentrosdecusto:
        ncentrodecusto = r[0]
        descricaodocentrodecusto = r[1]
        centrosdecusto[ncentrodecusto] = descricaodocentrodecusto
        
    
    return centrosdecusto

# PAGA A FORNECEDOR E RECEBE DE CLIENTE
def obter_uf_cliente(indice,interfacecontaspagarreceber):
    
    cursor = conexao.cursor()
    clicodigo = []
    clicodigodict = {}
    
    for l in indice:
        
        i = interfacecontaspagarreceber[l]
        
        codigocli = i['CodigoClienteFornecedor']
        clicodigo.append(codigocli)
    
    clicodigo = tuple(clicodigo)    
    
    #print('Clicodigo')
    #print(clicodigo)

    sql = 'TRUNCATE TABLE TBL_TMP_CODCLI_UF'
    cursor.execute(sql)

    for c in clicodigo:
        sql = f"INSERT INTO TBL_TMP_CODCLI_UF (CLI_CODIGO,UF) SELECT CLI_CODIGO,CLI_UF as UF FROM MXM_PRD.cliente_cli WHERE cli_codigo = '{c}'"
        cursor.execute(sql)
    
    sql = 'commit'
    cursor.execute(sql)

    #sql = f'SELECT cli_codigo,cli_uf as uf FROM MXM_PRD.cliente_cli WHERE cli_codigo IN {clicodigo}'

    sql = f'SELECT cli_codigo,uf FROM tbl_tmp_codcli_uf'
    
    uf_cursor = cursor.execute(sql)
    uftuplelist = uf_cursor.fetchall()

    for p in uftuplelist:
        clicodigodict[p[0]] = p[1]            
    
    return clicodigodict
   

def consultar_centro_custo(numerodocentrodecusto):        
    
    if len(numerodocentrodecusto) != 0:       
        descricaodocentrodecusto = centrosdecusto[numerodocentrodecusto]                
        centrodecusto = numerodocentrodecusto + ' - ' + descricaodocentrodecusto

    else:
        centrodecusto = ''
    
    return centrodecusto

def consultar_titulo_receber(empresa, mes, ano):
    
    #mes = '01' # FORÇADO PARA O SNG ANO INTEIRO
    
    datapgtoinicial = '01/'+mes+'/'+ano
    #print('datapgtoinicial: ',datapgtoinicial)

    # Subtrai um dia para obter o último dia do mês anterior
    if int(mes) != 12:
        ultimo_dia = date(int(ano), int(mes)+1, 1) - timedelta(days=1)
        
    else:
        ultimo_dia = date(int(ano)+1, 1, 1) - timedelta(days=1)

    
    datapgtofinal = str(ultimo_dia.day)+'/'+mes+'/'+ano
    #datapgtofinal = '31/12/'+ano # FORÇADO PARA O SNG ANO INTEIRO
    
    #print('datapgtofinal: ', datapgtofinal)
    
    response = post(f'{BASE_URL}/InterfacedoContasPagarReceber/ConsultaTituloReceber', verify=False, json={ # A MÁQUINA DO DATASTAGE PRECISA TER ACESSO A ESSA REDE.
                "AutheticationToken": {
                "Username": loginapi, 
                "Password": senhaapi,
                "EnvironmentName": ambiente
            },
            "Data": {
                "EmpresaEmitente": empresa,
                "DataRecebimentoInicial": datapgtoinicial,  # DATA DE PAGAMENTO/VENCIMENTO INICIAL. MÊS SEGUINTE AO DA COMPETÊNCIA
                "DataRecebimentoFinal": datapgtofinal  # DATA DE PAGAMENTO/VENCIMENTO FINAL. MÊS SEGUINTE AO DA COMPETÊNCIA
            }
        }                            )
    
    return response.json()

def conexao_bd(senha, usuario, host, sid):
    
    oracledb.init_oracle_client() 
    
    dsn = oracledb.makedsn(host=host, port=1521, sid=sid)
        
    connection = oracledb.connect(
        user=usuario,
        password=senha,
        dsn=dsn
    )

    print("Successfully connected to Oracle Database")

    return connection

def criar_dataframe(tituloareceber):

    dftituloareceber = DataFrame(tituloareceber)
    dftituloareceber = dftituloareceber.sort_values(by=['codigocliente','titulo']).reset_index(drop=True)
    
    #print(dftituloareceber)    
    
    return dftituloareceber

def filtrar_fluxocaixa(interfacecontaspagarreceber): # INDICES DOS DICIONÁRIOS COM O FLUXO DE CAIXA

    indice = []

    # PARA REQUEST COMPLETO        
    for t in interfacecontaspagarreceber:

        # LISTA DE DICIONÁRIOS COM AS CONTAS DE FLUXO DE CAIXA
        interfacegrupopagarreceber = t['InterfaceGrupoPagarReceber']
        
        for c in interfacegrupopagarreceber:
            pcontadofluxodecaixa = c['ContadoFluxodeCaixa']

            if pcontadofluxodecaixa == contadofluxodecaixa:
                indice.append(interfacecontaspagarreceber.index(t))
                break

    print('Total de Titulos: ', len(indice))
    
    if (len(indice)) == 0:
        raise NotitErro
    
    else:
        return indice
    
    
def jsonparsing(json):

    tempo_inicial = (time())  # em segundos
    
    centrodecusto = ''
    
    tituloareceber = []
            
    interfacecontaspagarreceber = json['Data']['InterfacedoContasPagarReceber']
    
    indice = filtrar_fluxocaixa(interfacecontaspagarreceber)
    
    clicodigodict = obter_uf_cliente(indice,interfacecontaspagarreceber)
    centrosdecusto = obter_centros_custo(indice,interfacecontaspagarreceber)
    
    #print(centrosdecusto)
       
    #print('Clicodigodict: ',clicodigodict)
    
    k = 0
    l = 0
    
    for l in indice:
        k = k+1
        
        i = interfacecontaspagarreceber[l]
                
        # SEÇÃO TÍTULO A RECEBER
        codigoclientefornecedor = i['CodigoClienteFornecedor']
        descricaodoclientefornecedor = i['DescricaodoClienteFornecedor']
        numerodotitulo = i['NumerodoTitulo']
        documentofiscal = i['DocumentoFiscal']
        descricaodostatusdotitulo = i['DescricaodoStatusdoTitulo']
        descricaodaempresaemitente = i['DescricaodaEmpresaEmitente']
        descricaodafilial = i['DescricaodaFilial']
        descricaodaempresarecebedora = i['DescricaodaEmpresaRecebedora']
        tipodetitulo = i['TipodeTitulo']
        descricaodotipodetitulo = i['DescricaodoTipodeTitulo']
        tipodecobranca = i['TipodeCobranca']
        descricaodotipodecobranca = i['DescricaodoTipodeCobranca']
        pedido = i['Pedido']
        datadeemissao = datetime.strptime(
            i['DatadeEmissao'], '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
        datadevencimento = datetime.strptime(
            i['DatadeVencimento'], '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
        datadecompetencia = datetime.strptime(
            i['DatadeCompetencia'], '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
        observacao = i['Observacao']
        valordotitulo = formatar_valor(i['ValordoTitulo'])
        valorcorrigido = formatar_valor(i['ValorCorrigido'])
        
        valordesconto = formatar_valor(i['ValordeDesconto']) 
            
        if len(str(i['DatadoDesconto'])) != 0:
            datadesconto = datetime.strptime(i['DatadoDesconto'], '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y'),
        else:
            datadesconto = i['DatadoDesconto']
            
        valordemulta = formatar_valor(i['ValordeMulta'])
        valordejuros = formatar_valor(i['ValordeJuros'])
                
        # SEÇÃO PAGAMENTO
        contadepagamento = i['ContadePagamento']
        agencia = i['Agencia']
        nomebancopagamento = i['NomeBancoPagamento']
        documentodepagamento = i['DocumentodePagamento']
        datadepagamento = datetime.strptime(
            i['DatadePagamento'], '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
        valorpago = formatar_valor(i['ValorPago'])        
        
        vnumerodotitulo = numerodotitulo
        
        # TRATAR OS CASOS DE FATURAS NÃO NUMERICAS, EX: 247904A PARA SER PELA TABELA DE TITULOS A RECEBER
        # Titulo = 247904A
        
        # TABELA DE ITENS DE FATURAMENTO. AQUI NÃO EXISTEM TÍTULOS COM LETRA EM SEU CÓDIGO
        sql = "select ifat_cdfatura as fatura, ifat_descricao as descricao, ifat_quantidade as quantidade, ifat_precoinf as precounitario, ifat_tpoper as tipooperacao, ifat_noccusto as centrodecusto, ifat_vlrprod as valortotal, ifat_vlrdecr as valordecrescimo, ifat_vlracres as valoracrescimo from MXM_PRD.itfatura_ifat ifat where ifat_cdfatura = :numerodotitulo"
                
        
        if vnumerodotitulo.isdigit():
            # print(str(k) + ' Titulo: ',vnumerodotitulo)
            
            cursor = conexao.cursor()
            cursor = cursor.execute(sql,numerodotitulo = vnumerodotitulo)
            tupla = cursor.fetchall()
                        
        elif vnumerodotitulo[0:len(vnumerodotitulo) - 1].isdigit():
                        
            vnumerodotitulo = vnumerodotitulo[0:len(numerodotitulo) - 1]
            #print(str(k) + ' Titulo: ',vnumerodotitulo)
            
            cursor = conexao.cursor()
            cursor = cursor.execute(sql,numerodotitulo = vnumerodotitulo)
            tupla = cursor.fetchall()
            
        else: # IGNORAR NO FATURAMENTO, TÍTULO DO CONTAS A RECEBER QUE TEM LETRA SEM SER NA ÚLTIMA POSIÇÃO
            continue       
        
        # SEÇÃO APROPRIAÇÃO (PRODUTOS). MAIS DE UMA OCORRÊNCIA POR TÍTULO. PEGANDO DO FATURAMENTO
        
        for g in tupla:                               
                
                valordogrupo = formatar_valor(g[6])
                valordecrescimo = formatar_valor(g[7])
                valoracrescimo = formatar_valor(g[8])
                centrodecusto = consultar_centro_custo(g[5])
                item = g[1]
                quantidade = formatar_quantidade(g[2])
                valorunitario = formatar_valor(g[3])           

                for j in i['InterfaceGrupoPagarReceber']:
                    
                    if j['NumerodoCentrodeCusto'] == centrodecusto:
                        break                    
                    
                codigodogrupo = j['CodigodoGrupo']
                descricaodogrupo = j['DescricaodoGrupo']
                contadofluxodecaixa = j['ContadoFluxodeCaixa']  
                                                                
                # DE CENTRO DE CUSTO PARA BAIXO, PEGAR DO FATURAMENTO  
                #print('K: ',k,' Tamanho da lista tuple: ',len(uftuplelist))
                                                 
                titulo = {
                                'titulo': vnumerodotitulo,
                                'documentofiscal': documentofiscal,
                                'codigocliente': codigoclientefornecedor,
                                'cliente': descricaodoclientefornecedor,
                                'uf' : clicodigodict[codigoclientefornecedor],
                                'statusdotitulo': descricaodostatusdotitulo,
                                'empresaemitente': descricaodaempresaemitente,
                                'filial': descricaodafilial,
                                'empresarecebedora': descricaodaempresarecebedora,
                                'pedidofaturamento' : pedido,
                                'tipodetitulo': tipodetitulo +' - '+ descricaodotipodetitulo,
                                'tipodecobranca': tipodecobranca +' - '+ descricaodotipodecobranca,
                                'emissao': datadeemissao,
                                'vencimento': datadevencimento,
                                'competencia': datadecompetencia,
                                'observacao': observacao,
                                'valordotitulo': valordotitulo,
                                'valorcorrigido': valorcorrigido,
                                'valordemulta': valordemulta,
                                'valordejuros': valordejuros,
                                'contadepagamento': contadepagamento,
                                'agencia': agencia,
                                'valordesconto' : valordesconto,
                                'datadesconto' : datadesconto,
                                'nomebancopagamento': nomebancopagamento,
                                'documentodepagamento': documentodepagamento,
                                'pagamento': datadepagamento,
                                'valorpago': valorpago,
                                'codigodogrupo': codigodogrupo,
                                'grupoderecebimento': descricaodogrupo,
                                'contadofluxodecaixa': contadofluxodecaixa,                    
                                'centrodecusto': centrodecusto,
                                'valorcentrocusto': valordogrupo,
                                'valordecrescimo' : valordecrescimo,
                                'valoracrescimo' : valoracrescimo,
                                'quantidade' : quantidade,
                                'valorunitario' : valorunitario,
                                'descricaoitem' : item                                                
                        }            

                tituloareceber.append(titulo)                                   
    
    tempo_final = (time())  # em segundos
    print('Tempo jsonparsing: ', tempo_final - tempo_inicial)
            
    return tituloareceber  

def validar_param_entrada(empresa,mes,ano,contadofluxodecaixa):
    
    mesatual = date.today().month
    anoatual = date.today().year
          
    #if int(ano) != anoatual:
    #    raise Anoinvalido    
    
    #elif int(mes) >= mesatual:
    #    raise Mesinvalido
    
    #elif len(empresa) != 3:
    if len(empresa) != 3:
        raise EmpErro
                            
    # Testa contadofluxodecaixa
    elif len(contadofluxodecaixa) < 6 or contadofluxodecaixa.isdigit() == False:
        raise FlxErro
            
    # Testa mês    
    elif len(mes) != 2 or mes is None or mes.isdigit() == False:
        raise MesErro

    # Testa ano
    elif len(ano) != 4 or ano is None or ano.isdigit() == False:
        raise AnoErro 

# MAIN
#
# NumeroTitulo: 252224, Pedido de Faturamento: 262684 

def criar_excel(dftituloareceber):
    
     # Limpando diretório com os arquivos antigos
    remove_arquivos_antigos('consulta receita sng.*')
     
    #nomeplanilha = 'consulta receita sng siseg.xlsx'
    nomeplanilha = f'consulta receita sng_{anopgto}{mespgto}.xlsx'
    
    # 16. Coluna Observação = 16
    #dftituloareceber.insert(loc=1, column='cod_cadu') # 'cadu' -> Nomenclatura da B3
    
    dftituloareceber.rename(columns={
                                        'observacao':'cod_cadu',
                                        'uf' : 'uf_cliente'
                                    },inplace=True)
    
    listacodcadu = []

    for observacao in dftituloareceber['cod_cadu']:  
        if len(re.findall(r'\d+',observacao)) != 0:
            #print('Observacao: ',observacao)               
            result = re.findall(r'\d+',observacao)[0]
            listacodcadu.append(result)
        
        else: # EVITANDO O ERRO Length of values does not match length of index
            listacodcadu.append(None)
    
        
    dftituloareceber['cod_cadu'] = listacodcadu
    
    dftituloarecebercompleto = dftituloareceber
    
    dftituloareceber = dftituloareceber[['titulo','documentofiscal','codigocliente','cliente','uf_cliente','statusdotitulo','empresaemitente','filial','empresarecebedora','pedidofaturamento',
                                        'tipodetitulo','tipodecobranca','emissao','vencimento','competencia','cod_cadu','contadepagamento','nomebancopagamento',
                                        'documentodepagamento','pagamento','valordemulta','valordejuros','codigodogrupo','grupoderecebimento','contadofluxodecaixa','descricaoitem']].drop_duplicates()
    
    dftituloareceber = dftituloareceber.sort_values(by=['competencia','titulo'])        
        
    dftituloareceber.to_excel(nomeplanilha)
    wb = load_workbook(nomeplanilha)
    ws = wb.active
    ws.delete_cols(1)
            
    ajusta_largura(ws)
    wb.save(nomeplanilha)
    
    return nomeplanilha

# DAQUI PARA CIMA, É A RESPEITO DA IMPLEMENTAÇÃO PARA CONSUMIR A API
def conexao_sftp():
    
    cnopts = sftp.CnOpts() # Para não utilizar chave
    cnopts.hostkeys = None # Para não utilizar chave
    s = sftp.Connection(ip, username=username, password=password, cnopts = cnopts) 
    
    return s

def cria_planilha_cob():
    
    global s                
    s = conexao_sftp()                                
    s.chdir('./Projetos/SNG/Dados_Auxiliares')   # temporarily chdir to bifiles
    
    planilhacobnome = f'cob_fenaseg_{anopgto}.xlsx'
        
    obtem_planilhas_cob() # TRAZENDO AS PLANILHAS PARA A MÁQUINA LOCAL, OU SEJA, DATASTAGE  
        
    planilhas = []
    
    files = os.listdir(diretoriolocalplanilhassng) # CARREGAS AS PLANILHAS DESSE DIRETÓRIO SEM FAZER OUTRAS VERIFICAÇÕES
    
    # LOCALIZA PLANILHAS
    for file in files:
        #print('Arquivo: ',file)
        
        if re.match(planilhassng,file):
            #print('Arquivo: ',file)            
            planilhas.append(file)
    
    # CONCATENA DATAFRAMES A PARTIR DAS PLANILHAS
    if len(planilhas) > 0:
        print('Lendo planilha: ', planilhas[0])
        
        #EDITANDO PLANLHA COB - DELETA A PRIMEIRA LINNHA E MUDA A DESCRIÇÃO DAS CÉLULAS A1 E I1
        planilha = diretoriolocalplanilhassng+planilhas[0]
        wb = load_workbook(planilha)
        ws = wb.active
        
        
        ws.delete_rows(1)
        ws['A1'] = 'cod_cliente'
        ws['I1'] = 'desconto_repasse'
        wb.save(planilha)
                
        dfsng = cria_competencia(planilha) # DEVOLVE UM DATAFRAME COM A COMPETÊNCIA CRIADA
        
        # CONCATENA AS PLANILHAS RESTANTES
        for i in range(len(planilhas)-1):
            print('Lendo planilha: ', planilhas[i+1])
            
            #EDITANDO PLANLHA COB
            planilha = diretoriolocalplanilhassng+planilhas[i+1]
            wb = load_workbook(planilha)
            ws = wb.active
            ws.delete_rows(1)
            ws['A1'] = 'cod_cliente'
            ws['I1'] = 'desconto_repasse'    
            wb.save(planilha)
                         
            dfsng1 = cria_competencia(planilha)            
            dfsng = concat([dfsng,dfsng1],ignore_index=True)              
        
        dfsng = dfsng.sort_values(by="competencia")
        
        # FORMATANDO A COLUNA "COMPETENCIA" AQUI, PARA O FORMATO DD/MM/YYYY, PARA QUE A ORDENAÇÃO ANTERIOR PUDESSE SER FEITA.
        dfsng['competencia'] = to_datetime(dfsng['competencia']).dt.strftime('%d/%m/%Y')
        
        # REMOVENDO ARQUIVOS ANTIGOS
        remove_arquivos_antigos('cob_fenaseg.*')
                
        dfsng.to_excel(planilhacobnome)
        
        wb = load_workbook('./'+planilhacobnome)
        ws = wb.active
        ws.delete_cols(1)
                        
        ajusta_largura(ws)        
        wb.save('./'+planilhacobnome)        
                        
        s.close()
        
        #print(dfsng)
        
        #try: # SE NÃO EXISTIR A COMPETENCIA PASSADA NO DF, PARAR O PROCESSO E ALERTAR SOBRE OS ARGUMENTOS PASSADOS
            
        dfsng = read_excel(planilhacobnome, dtype=str) # SÓ PASSANDO TUDO PARA STRING, PARA FUNCIONAR A TROCA DE '.' POR ',' EM desconto_repasse e fenaseg
        
        inserir_cob_bd(dfsng)
           
    else:
        print('Náo foram encontradas as planilhas Dcobranca Fenaseg.xlsx')
    
    s.close()

def obtem_planilhas_cob():
    
    # Limpando diretório local
    files = os.listdir(diretoriolocalplanilhassng)
    
    for f in files:
        filepath = diretoriolocalplanilhassng+f
        if os.path.isfile(filepath):
            print('Removendo planilhas existentes. Arquivo: ', filepath)
            os.remove(filepath)
    
           
    # DATASTAGE obtendo todos os arquivos do diretório especificado do servidor do QlikSense
    
    s.get_d('./Faturamento SNG',diretoriolocalplanilhassng)
    

def cria_competencia(planilha): # DIRETÓRIO + PLANILHA
    
    wb = load_workbook(planilha)
    aba = wb.sheetnames[0]
    anomes = aba
    
    #print('Anomes: ',anomes)
    
    ano = anomes[0:4]
    mes = anomes[4:]

    if int(mes) != 12:
        ultimo_dia = date(int(ano), int(mes)+1, 1) - timedelta(days=1) # COMPETÊNCIA

    else:
        ultimo_dia = date(int(ano)+1, 1, 1) - timedelta(days=1) # COMPETÊNCIA
        
    
    competencia = ultimo_dia    
        
    df = read_excel(planilha) # PRIMEIRA VEZ EM QUE SE ABRE UM DATAFRAME
    df['competencia'] = competencia        
    
    return df


def remove_arquivos_antigos(padrao):
    files = os.listdir('./')
    for file in files:
        if os.path.isfile(file):
            if len(re.findall(rf"{padrao}",file)) != 0:
                fileremove = re.findall(rf"{padrao}",file)[0]
                print('Arquivo a ser removido: ',fileremove)
                os.remove(f'./{fileremove}') 


# ESCREVE NA TABELA TBL_TMP_COB_FENASEG
def inserir_cob_bd(df):
     
    ano = anopgto

    df['desconto_repasse'] = df['desconto_repasse'].str.replace('.',',')
    df['fenaseg'] = df['fenaseg'].str.replace('.',',')
     
    dfcompetencias = df['competencia'].drop_duplicates()
    listacompetenciasdf = list(dfcompetencias) # LISTA DE STRINGS
    print("CompetenciasB3planilhas: ",listacompetenciasdf)
        
    dfcompetenciastbl = lista_tabela("SELECT DISTINCT competencia FROM TBL_COB_FENASEG")['COMPETENCIA']
    dfcompetenciastbl = to_datetime(dfcompetenciastbl, dayfirst=True).sort_values().dt.strftime('%d/%m/%Y') # TRANSFORMA EM STRING
    listacompetenciastbl = []
    for i in dfcompetenciastbl:
        listacompetenciastbl.append(i)
        
    print("CompetenciasB3tbl: ",listacompetenciastbl)
    
    if len(listacompetenciastbl) != 0:
        #print('Tabela cob com registros')
        listaanocompetenciastbl = []
         
        for r in listacompetenciastbl:
            anocompetenciatbl = r[6:] # POSSUI DUPLICATAS DE ANO, POR ISSO O TRATAMENTO ABAIXO
            if anocompetenciatbl not in listaanocompetenciastbl:
                listaanocompetenciastbl.append(anocompetenciatbl)
        
        print("AnoCompetenciasB3tbl: ",listaanocompetenciastbl)
        
        if ano not in listaanocompetenciastbl: # PREPARANDO A TABELA DE COBRANÇAS FENASEG PARA UM NOVO ANO.
            print('Truncando a tabela cob. Novo ano: ',ano)
            conexao.cursor().execute('TRUNCATE TABLE TBL_COB_FENASEG')
            
            print('Tabela cob sem registros. Fazendo inserção')
            engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
            df.to_sql('tbl_cob_fenaseg', engine, if_exists='append', index=False)
        
        else:
        
            # VERIFICANDO EXISTÊNCIA DE COMPETENCIA NO BD
            for competencia in listacompetenciasdf:
                #print (type(competencia))
                #print(type(listacompetenciastbl))
                if competencia in listacompetenciastbl:
                    print(f'Competência {competencia} da planilha da B3 já existente no BD. Não haverá inserção de novos registros na tabela TBL_COB_FENASEG.')
                    #break # SE EXISTE UMA COMPETENCIA DO DF NO BD, IGNORA TODO O DF, POIS ELE SEMPRE É DISPONIBILIZADO NO MÊS SEGUINTE, COM DADOS DO MÊS ANTERIOR.
                
                else:
                    print(f'Inserindo nova competência {competencia} na tabela TBL_COB_FENASEG')
                    engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
                    df.loc[df['competencia'] == competencia].to_sql('tbl_cob_fenaseg', engine, if_exists='append', index=False)
                    #break
            
    else:
        print('Tabela cob sem registros. Fazendo inserção')
        engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
        df.to_sql('tbl_cob_fenaseg', engine, if_exists='append', index=False)
        

def lista_tabela(sql):
    
    cursor = conexao.cursor()
    #cursor.execute("ALTER SESSION SET nls_date_format = 'DD/MM/YY'")
    cursor.execute("ALTER SESSION SET NLS_NUMERIC_CHARACTERS = ',.'")    
    
    print(sql)
    
    tabela_cursor = cursor.execute(sql)
    tabela = tabela_cursor.fetchall()  # LISTA DE TUPLAS
            
    return cria_dataframe(tabela, cursor) # DEVOLVE UM DATAFRAME, AONDE AS COLUNAS, SÃO AS COLUNAS DO BD


def cria_receita_sng(planilhareceitasng):

    dfreceitasng = read_excel(planilhareceitasng) # LÊ A PLANLHA DE RECEITAS
    dfreceitasngpago = dfreceitasng.loc[dfreceitasng['statusdotitulo'] == 'Pago']

    insere_receita_bd(dfreceitasngpago)
    

# AQUI É VERIFICADA SE EXISTE A COMPETÊNCIA NO BD, CASO NÃO, UMA EXCEÇÃO É LANÇADA. 
# ESCREVE NA TABELA TBL_TMP_RECEITA_SNG
def insere_receita_bd(dfreceitasngpago):
    
    cursor = conexao.cursor()
    
    ano = anopgto
    
    #print('Tipo anopgto: ', type(anopgto))
    
    verifica_competencia(cursor)
    
    dfmespagamento = to_datetime(dfreceitasngpago['pagamento'],dayfirst=True).dt.month
    mespagamento = dfmespagamento.iloc[0]

    #print('Mes pagamento: ',mespagamento)
                
    # SUPONDO QUE ESTAMOS NO MÊS DE DEZEMBRO
    # NÃO É PARA PERMITIR EXECUTAR CONSULTA DE RECEITA PARA DEZENBRO. ESSA EXCEÇÃO ESTÁ NA VALIDAÇÃO DOS PARÂMETROS DE ENTRADA.
    # SE O MÊS NÃO ACABOU AINDA, OU SE NÃO TEM O ARQUIVO DA B3 DE COMPETÊNCIA OUTUBRO PARA OS PAGAMENTOS DE NOVEMBRO
    
    listaanopagamentostbl = []
            
    sql = f"SELECT distinct pagamento FROM TBL_TMP_RECEITA_SNG ORDER BY to_date(pagamento,'DD/MM/RRRR')"
    pagamentostbl = cursor.execute(sql).fetchall() # NÃO TEMOS EXCEÇÃO AQUI NO CASO DE NÃO HAVER REGISTROS.
        
    if len(pagamentostbl) !=0: 
        for tupla in pagamentostbl:
            listaanopagamentostbl.append(tupla[0][6:])
        
        listaanopagamentostbl = list(set(listaanopagamentostbl))
        print('Anos existentes na tabela de receita: ', listaanopagamentostbl)
    
        # SE O NOVO ANO NÃO EXISTE, LIMPAR A TABELA DE RECEITA.
        if ano not in listaanopagamentostbl:
            print("Limpando tabela de receita para novo ano")
            cursor.execute("TRUNCATE TABLE TBL_TMP_RECEITA_SNG")
        
            print('Tabela de receita vazia. Inserindo registros')        
            engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
            dfreceitasngpago.to_sql('tbl_tmp_receita_sng', engine, if_exists='append', index=False)        
        
        else:
            listamespagamentostbl = []
        
            #  VERIFICANDO SE MÊS DE PAGAMENTO JÁ EXISTE NA TABELA DE RECEITA            
            for tupla in pagamentostbl:
                listamespagamentostbl.append(tupla[0][3:5])
            
            listamespagamentostbl = list(set(listamespagamentostbl)) # RETIRANDO DUPLICATAS DA LISTA, MAS O SET DEVOLVE UM DICIONÁRIO, POR ISSO O list
                                
            print('Meses de pagamento existentes na tabela de receita: ',sorted(listamespagamentostbl))
                    
            if (str(mespagamento).zfill(2) in listamespagamentostbl):
                print(f'Mês de pagamento {str(mespagamento).zfill(2)} já existente no BD. Não haverá inserção de novos registros na tabela de receita')
                
            else:
                print(f'Inserindo novo mês {str(mespagamento).zfill(2)} na tabela de receita')
                engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
                dfreceitasngpago.to_sql('tbl_tmp_receita_sng', engine, if_exists='append', index=False)
    
                               
    else:
        print('Tabela de receita vazia. Inserindo registros')        
        engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
        dfreceitasngpago.to_sql('tbl_tmp_receita_sng', engine, if_exists='append', index=False)
        
    cursor.close()


def verifica_competencia(cursor):
    
    # VERIFICANDO A EXISTÊNCIA DA COMPETÊNCIA NO BD. TABELA DE COBRANÇA.
    print('Verificando a existência da competência no BD') 
        
    mescompetencia = int(mespgto) - 1
    if mescompetencia == 0: # VERIFICAR ESSE IF PARA O CASO DE JANEIRO DO ANO SEGUINTE.
            mescompetencia = 12
            anocompetencia = int(anopgto) - 1
    else:
            anocompetencia = anopgto

    mescompetencia = str(mescompetencia).zfill(2)
    
    print('Mês e ano de Competência para mes e ano de pgto informados: ',f'{mescompetencia}/{str(anocompetencia)}')

    sql = f"SELECT DISTINCT COMPETENCIA FROM TBL_COB_FENASEG WHERE COMPETENCIA LIKE '%/{mescompetencia}/{anocompetencia}'"
    retorno = list(cursor.execute(sql).fetchall())[0][0] # SE A COMPETENCIA NÃO EXISTIR NA TABELA DE COBRANÇA, SERÁ LANÇADA A EXCEÇÃO IndexError
    mescompetencia = retorno[3:5]
    print('Mês e ano de Competência - Retorno BD: ',mescompetencia+'/'+str(anocompetencia))
        
    # SE A COMPETÊNCIA EXISTE NO BD, TABELA DE COBRANÇA, A EXECUÇÃO CONTINUA E NÃO OCORRE EXCEÇÃO, PROSSEGUINDO COM A CRIAÇÃO DA RECEITA.

    #print('Tipo mescompetencia: ', type(mescompetencia))
    #print('Tipo mespgto: ', type(mespgto))
    #print('Mescompetencia: ', mescompetencia)
    #print('Mespgto: ', mespgto)

    print('Já existe a competência no BD')
    

def cria_flxcx_sng():
    
    ano = anopgto # PARA A DATA_FLXCX
    mes = mespgto # PARA A DATA_FLXCX
        
    cursor = conexao.cursor()
    cursor.execute("ALTER SESSION SET nls_date_format = 'DD/MM/YY'")
    
    dfdatasflxcxtbl = lista_tabela("SELECT DISTINCT data_flxcx FROM TBL_TMP_FLUXO_CAIXA_SNG").sort_values(by='DATA_FLXCX')['DATA_FLXCX']
    dfmesesflxcxtbl = to_datetime(dfdatasflxcxtbl).dt.month.drop_duplicates()
    listamesesflxcxtbl = []
    
    dfanoflxcxtbl = to_datetime(dfdatasflxcxtbl).dt.year.drop_duplicates()
        
    for i in dfmesesflxcxtbl:
        listamesesflxcxtbl.append(str(i).zfill(2))
        
    print("Meses Flxcxtbl: ",listamesesflxcxtbl)
    
    if len(listamesesflxcxtbl) != 0:
        #print('Tabela cob com registros')
        anoflxcxtbl = dfanoflxcxtbl.iloc[0] 

        print("AnoFlxcxtbl: ",anoflxcxtbl)
        
        if ano != str(anoflxcxtbl): # PREPARANDO A TABELA DE COBRANÇAS FENASEG PARA UM NOVO ANO.
            print('Truncando a tabela TBL_TMP_FLUXO_CAIXA_SNG. Novo ano: ',ano)
            conexao.cursor().execute('TRUNCATE TABLE TBL_TMP_FLUXO_CAIXA_SNG')
                
            print('Tabela de fluxo caixa sng vazia. Fazendo inserção')
            cursor.execute(f'begin PKG_BI_CORPORATIVO_SNG.PRC_EXECUTA_JOBS_SNG(pano => {anopgto}, pmes => {mespgto}); end;')
            cursor.close() 
            
        else:
            # VERIFICANDO EXISTÊNCIA DE MES NO BD
            if mes in listamesesflxcxtbl: # VERIFICAR ESSE CONDICIONAL
                print(f'Mes de pagamento {mes} já existente no BD. Não haverá inserção de novos registros na tabela TBL_TMP_FLUXO_CAIXA_SNG.')
                            
            else:
                print(f'Inserindo novo mes {mes} na tabela TBL_TMP_FLUXO_CAIXA_SNG')
                cursor.execute(f'begin PKG_BI_CORPORATIVO_SNG.PRC_EXECUTA_JOBS_SNG(pano => {anopgto}, pmes => {mespgto}); end;')
                cursor.close()
                        
    else:
        print('Tabela de fluxo caixa sng vazia. Fazendo inserção')
        cursor.execute(f'begin PKG_BI_CORPORATIVO_SNG.PRC_EXECUTA_JOBS_SNG(pano => {anopgto}, pmes => {mespgto}); end;')
        cursor.close()         
    

def cria_sng():
    
    global s                
    s = conexao_sftp()                                
    s.chdir('./Projetos/SNG/Dados_Auxiliares')   
            
    nome = f'sng_{anopgto}.csv' # ÚNICO ARQUIVO QUE FICA. NÃO É APAGADO NO FINAL DO PROCESSO.
    
    dataframe = lista_tabela('SELECT * FROM V_RECEITA_SNG_ESTADO')
    
    cria_arquivo(dataframe,nome)
    envia_arquivo(nome)
    
    # CRIANDO UM EXCEL TAMBÉM        
    dataframe.to_excel(f'sng_{anopgto}.xlsx')
    wb = load_workbook(f'sng_{anopgto}.xlsx')
    ws = wb.active
    ws.delete_cols(1)
    ajusta_largura(ws)
    wb.save(f'sng_{anopgto}.xlsx')   
    
    # PARA O ARQUIVO DE JUROS
    nome = f'sng_{anopgto}_juros.csv' # ÚNICO ARQUIVO QUE FICA. NÃO É APAGADO NO FINAL DO PROCESSO.
    
    dataframe = lista_tabela('SELECT * FROM V_RECEITA_SNG_ESTADO_JUROS')
    
    cria_arquivo(dataframe,nome)
    envia_arquivo(nome)
    
    # CRIANDO UM EXCEL TAMBÉM        
    dataframe.to_excel(f'sng_{anopgto}_juros.xlsx')
    wb = load_workbook(f'sng_{anopgto}_juros.xlsx')
    ws = wb.active
    ws.delete_cols(1)
    ajusta_largura(ws)
    wb.save(f'sng_{anopgto}_juros.xlsx')   


def cria_dataframe(resultado, cursor):
    
    #print('Resultado: ',resultado)
    df = DataFrame(resultado, columns=[desc[0] for desc in cursor.description])    
    cursor.close()

    #print('Dataframe: ',df)

    return df
    
  
def ajusta_largura(worksheet):
    
    for col in worksheet.columns:
        #print('Col: ',col)
        max_length = 0
        for cell in col:
            #print('Cel: ',cell)
            if isinstance(cell.value, str):
                max_length = max(max_length, len(str(cell.value))) # RETORNA O MAIOR TAMANHO ENTRE 2 TAMANHOS
            elif isinstance(cell.value, int) or isinstance(cell.value, float):
                max_length = max(max_length, len(str(cell.value)))
        
        col = col[0].column_letter
        #print('1 col da tupla: ',col)
        worksheet.column_dimensions[col].width = max_length + 4

    # writer.close()
    
    return worksheet


def cria_arquivo(dataframe,nome):
    
    hoje = date.today()
    
    data_formatada = hoje.strftime("%d%m%Y")    
   
    #dataframe.to_csv(nome, sep='|', index=False, decimal=',')
    dataframe.to_csv(nome, sep='|', index=False)

def envia_arquivo(nome):
        
        print(f'Enviando {nome}')
        s.put(nome) 
        
        # FOI NECESSÁRIO FAZER MODIFICAÇÕES NO CÓDIGO do pysftp, linha 848, código sftp_client.py     


# BANCO HMLG
#senhabd = 'MXM_HMLG#$01'
#usuariobd = 'MXM_HMLG'
#host = 'srv2033.cnseg.int'
#sid = 'HOMOLOG'

# BANCO PRD
senhabd = 'zvga2jd0871'
usuariobd = 'QLIK_MXM'
host = '10.10.20.61'
sid = 'MXMWEB2'

# AMBIENTE HMLG
#BASE_URL = 'https://mxm-hmlg.cnseg.org.br/mxmhom/api'
#loginapi = 
#senhaapi = 
#ambiente = 'WEBMANAGERHOM'

# AMBIENTE PRD
BASE_URL = 'https://mxm.cnseg.org.br/Producao/api'
#BASE_URL = 'https://10.10.40.80/Producao/api'
ambiente = 'PRODUCAO'
loginapi = 'INTEGRACAO_MXM'
senhaapi = 'Wd2&9ZL6We4!'

# ENVIO DOS ARQUIVOS
ip = '10.10.40.84'
username = 'qlik_admin'
password = '*wEd[<VTmb6uUql'

planilhassng = r"Dcobran[cç]a.*.xls.*"
diretoriolocalplanilhassng = 'G:\\Meu Drive\\Cursos e Treinamentos\\Cientista de Dados\\Treinamento Python\\Consulta receita despesa\\consulta_receita_despesa\\Scripts\\Planilhas SNG\\'
#diretoriolocalplanilhassng = 'E:\\DATASTAGE\\PROJETOS\\PRD\\QLIKVIEW\\EnviaCSV\\Scripts\\Planilhas SNG\\'
#diretoriolocalplanilhassng = 'C:\\Users\\qlik_admin\\Scripts\\consulta_receita_despesa\\Scripts\\Planilhas SNG\\'

def main():
 try:    
        global contadofluxodecaixa
        global mespgto
        global anopgto
        global conexao
        global ambiente
        global loginapi
        global senhaapi
                            
        #if (len(argv)) != 5:
        #    raise ParamErro
        
        #else:
                        
        contadofluxodecaixa = '030101'
        empresa = 'F01'
                    
        ano = date.today().year  # ANO ATUAL
        mes = date.today().month # MÊS ATUAL

        mespgto = mes - 1
        anopgto = ano

        if mespgto == 0: # PARA O ANO SEGUINTE
            mespgto = 12
            anopgto = ano - 1
        
        # CASO PRECISE FORÇAR A CARGA DE DADOS. UTILIZAR NÚMEROS INTEIROS
        #anopgto = 2024
        #mespgto = 12
        

        mespgto = str(mespgto).zfill(2)
        anopgto = str(anopgto) 

        print('Mes pagamento: ',mespgto)
        print('Ano pagamento: ',anopgto)

        #validar_param_entrada(empresa = argv[1], mes = argv[2], ano = argv[3], contadofluxodecaixa = argv[4])

        validar_param_entrada(empresa = empresa, mes = mespgto, ano = anopgto, contadofluxodecaixa = contadofluxodecaixa)

        conexao = conexao_bd(senhabd, usuariobd, host, sid) 
        
        json = consultar_titulo_receber(empresa = empresa, mes = mespgto, ano = anopgto) # OBTENDO JSON. LIBERAR ACESSO PARA A MÁQUINA DO DATASTAGE ACESSAR API DO MXM
        #print('JSON: ',json)
        
        # ÚLTIMA VALIDAÇÃO, APÓS A OBTENÇÃO DO JSON
        try:
            json['Message']                
            
        except KeyError as e: # NAO TEM A CHAVE NO DICIONÁRIO, ENTÃO TEM QUE TER A CHAVE 'Messages' PROSSEGUE EXECUÇÃO
            if len(json['Messages']) != 0:
                raise ErroAPI
            
        else: # VERIFICA SE TEM A CHAVE 'Message' no DICIONÁRIO. LANÇAR A EXCEÇÃO
            if len(json['Message']) != 0:
                raise ErroGeralAPI
                                                             

 except EmpErro:
    print("O código da empresa deve possuir 3 caracteres")
    
 except ParamErro:
    print("Número incorreto de parâmetros. Devem ser 4 Empresa Mêspagamento Anopagamento ContaFLuxodeCaixa (Ex:python.exe consulta_titulo_receber.py F01 01 2024 030101) ")

 except FlxErro:
    print("A conta do fluxo de caixa deve ter pelo menos 6 dígitos numéricos")   

 except MesErro:
    print("O mês deve ter dois dígitos numéricos")

 except AnoErro:
    print("O ano deve ter quatro dígitos numéricos")

 except ErroAPI:
    print('Erro da API: ', json['Messages'][0]['Message'])

 except ErroGeralAPI:
    print('Erro Geral API: ', json['Message'])
    
 except Mesinvalido:
    print("Mes de pagamento inválido. Mês ainda não terminado")
    
 except Anoinvalido:
    print("Ano inválido. O ano informado deve ser o atual")

 except oracledb.OperationalError as e:
    print('Erro de conexão ao Oracle')

# SE NÃO TIVER NENHUM ERRO, VEM PARA CÁ
 else:
     try:
        
        global centrosdecusto
        
        cria_planilha_cob() # CRIADA MENSALMENTE. COM DADOS DO MÊS ANTERIOR
        
        centrosdecusto = {} # DICIONÁRIO DE CENTROS DE CUSTO
        tituloareceber = jsonparsing(json) # O RETORNO É UMA LISTA DE DICTIONÁRIOS
        dftituloareceber = criar_dataframe(tituloareceber)  
        
        planilhareceitasng = criar_excel(dftituloareceber) # FORMATANDO O DATAFRAME. CRIADA MENSALMENTE COM DADOS DO MÊS ANTERIOR        
        
        cria_receita_sng(planilhareceitasng) # EXECUTA MESMO QUE NÃO EXISTAM AS PLANILHAS DE COBRANÇA FENASEG
              
        cria_flxcx_sng() # CRIADA MENSALMENTE COM DADOS DO MÊS ANTERIOR.
        
        cria_sng() # CRIADA MENSALMENTE COM DADOS DO MÊS ANTERIOR.
                                                 
        conexao.close()
    
        tempo_final = (time()) # em segundos
    
        print(f"Tempo Total {tempo_final - tempo_inicial} segundos")
        
     except NotitErro:
        print("Sem títulos para os argumentos passados")
    
     except IndexError:
         print("Competencia não existente no BD. Incluir planilha de competência da B3. REEXECUTAR O PROCESSO")
         
         remove_arquivos_antigos('cob_fenaseg.*')
    
        
if __name__ == "__main__":        
    main()