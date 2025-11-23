'''
    Ele irá gerar a planilha de edição manual, que será disponibilizada para os usuários nos diretórios 
    \\srv4036\\Planilha EdManual e \\srv4036\\Planilha EdManual - DRE

    Os arquivos abaixo, gerados pelo script Envia.csv, precisam estar disponíveis. 
    - v_dre_real_qlik_{ano}.csv'
    - v_fluxo_caixa_{ano}.csv

    Planilha protegida por senha. Ela será informada para um usuário de DRE (Claudia) e um usuário de Fluxo de Caixa (Shelden)
    Obs: Fazer o backup dos arquivos a cada novo mês. Senha Atual: cnseg246$
    
    O NOME DO ARQUIVO É UTILIZADO PARA O QLIK SABER QUAL É O ANOMES
    
    OS VALORES DA ABA EDIÇÃO MANUAL NÃO DEVEM SER EXCLUÍDOS, SENÃO, SERÃO SUBSTITUÍDOS PELOS VALORES DO MXM
    
    A CADA NOVO ANO, AS PLANILHAS DE ED MANUAL DEVE SER MOVIDAS PARA UMA PASTA DE BACKUP
'''

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from pandas import read_csv, read_excel
from pandas import ExcelWriter
from pandas import to_datetime
from datetime import date
import sqlalchemy as sa
from sqlalchemy.dialects.oracle import (
                                        CLOB, 
	                                    NUMBER, 
	                                    CLOB, 
	                                    CLOB
                                        )
import pysftp as sftp
import warnings
import re

warnings.filterwarnings('ignore','.*Failed to load.*')


class NaoAtualizaBD(Exception):
    pass

class ParamError(Exception):
    pass


def conexao_sftp():
    
    cnopts = sftp.CnOpts()  # Para não utilizar chave
    cnopts.hostkeys = None # Para não utilizar chave
    s = sftp.Connection(ip, username=username, password=password, cnopts = cnopts) 
    
    return s

def envia_arquivo(nome):
        
        print(f'Enviando {nome}')
        s.put(nome)  # upload file to public/ on remote
        
        # FOI NECESSÁRIO FAZER MODIFICAÇÕES NO CÓDIGO do pysftp, linha 848, código sftp_client.py
    
        
def inserir_no_bd(tabela,df,processo):
    
    print('Quantidade de elementos do DataFrame: ',df.size)
        
    engine = sa.create_engine(f'oracle+oracledb://{usuariobd}:{senhabd}@{host}:1521/{sid}',thick_mode=True)
    
    with engine.connect() as conn:
        query = sa.select(sa.func.count()).select_from(sa.text(tabela))
        result = conn.execute(query)
        
    if df.size == 0 and result.scalar() != 0:
        print('Planlha sem registros e BD com registros.')
        raise NaoAtualizaBD        
        
    elif df.size != 0:
        if processo == 'dre':
            dtype = {
                    "ID" : CLOB, 
                    "valor incrementado" : NUMBER(19,2), 
                    "Data Modificação" : CLOB, 
                    "Motivo" : CLOB
            }
        else:
            dtype = {
                    "CONTA" : CLOB, 
                    "valor incrementado" : NUMBER(19,2), 
                    "Data Modificação" : CLOB, 
                    "Motivo" : CLOB
            }
            
        df.to_sql(tabela,engine, if_exists='replace', index=False, dtype=dtype) # O replace faz um truncate antes
    

def ajusta_largura(worksheet):
        
    for col in worksheet.columns:
        #print('Col: ',col)
        max_length = 0
        for cell in col:
            #print('Cel: ',cell)
            if isinstance(cell.value, str):
                max_length = max(max_length, len(str(cell.value))) # RETORNA O MAIOR TAMANHO ENTRE 2 TAMANHOS
            elif isinstance(cell.value, int) or isinstance(cell.value, float):
                max_length = max(max_length, len(str(cell.value)))
        
        col = col[0].column_letter
        #print('1 col da tupla: ',col)
        worksheet.column_dimensions[col].width = max_length + 2

    # writer.close()
    
    return worksheet


def cria_planilha(df,planilha):
    
    #df[['Usuário']] = ''
    #df[['Data']] = ''    
    
    with ExcelWriter(planilha, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Referencia MXM') # CRIANDO NOVO EXCEL
                    
    workbook = load_workbook(planilha)
    workbook.create_sheet('Edicao Manual')
    worksheet = workbook['Edicao Manual']
                    
    worksheet = workbook['Referencia MXM']
    #worksheet.delete_rows(1)
    ajusta_largura(worksheet)
                     
    workbook.save(planilha)
    
    return workbook
    
   
def cria_planilha_dre(ano,mes):
    
    global s

    s = conexao_sftp()                                
    s.chdir('./Projetos/DRE/Dados_Auxiliares/Planilha EdManual')   # temporarily chdir to bifiles
    
    if mes == None and ano == None:
       ano = date.today().year
       mes = date.today().month 
               
       if mes == 1:
        backup_planedmanuais(s,ano)
        
        ano = ano - 1
        mes = 12
        
       else:
        mes = mes - 1 # OS DADOS SÃO DO MÊS ANTERIOR.
    
    elif mes is not None and ano is not None:
       mes = mes
       ano = ano
        
       #else:
       # mes = mes - 1
    
    else:
        raise ParamError
       
    print('Criando planilha de edição manual de DRE...')
    
    nomeplanilhaedmanual = f'dre_edmanual_{ano}{str(mes).zfill(2)}.xlsx' # ARQUIVO NOMEADO PARA O MES ANTERIOR.
    
    try:
        
        df = le_csv_dre(f'v_dre_real_qlik_{str(ano)}.csv',ano,mes) # REALIZA O FILTRO DOS DADOS DO MÊS ANTERIOR. LENDO ARQUIVO 
                                                                   # DA PASTA LOCAL DO SCRIPT
        
        #print('Lendo csv do mes anterior')
        #print(df)
        
        # SE A PLANLHA EXISTIR, CRIA NOVAMENTE, SÓ QUE COM OS DADOS DA EDIÇÃO MANUAL.
        #files = os.listdir('./')
        files = s.listdir('./')
        
        print('Arquivos no diretório de planilhas de edição manual: ',files)
            
        if nomeplanilhaedmanual in files:
            print('Planilha de edicao manual existente: ',nomeplanilhaedmanual)
            s.get(f'./{nomeplanilhaedmanual}')
            dfedmanual = le_aba_edmanual(nomeplanilhaedmanual)
            
            print('Aba Ed Manual')
            print(dfedmanual)
            
            # CRIA NOVAMENTE, SÓ QUE COM OS DADOS DA EDIÇÃO MANUAL.
            cria_planilha(df,nomeplanilhaedmanual) # CRIA AS ABAS REFERÊNCIA MXM E EDIÇÃO MANUAL, ESTA ÚLTIMA SEM DADOS
            
            with ExcelWriter(nomeplanilhaedmanual, engine='openpyxl',if_sheet_exists='replace',mode='a') as writer:
                dfedmanual.to_excel(writer, index=False, sheet_name='Edicao Manual') # INSERINDO DADOS DA EDIÇÃO MANUAL
            
            inserir_no_bd('tbl_tmp_dre_edmanual',dfedmanual, 'dre') # PARA EFEITO DE TESTE, DEVE-SE FAZER TRUNCATE NESSA TABELA ANTES
                    
        else:
            workbook = cria_planilha(df,nomeplanilhaedmanual) # CRIA AS ABAS REFERÊNCIA MXM E EDIÇÃO MANUAL
            worksheet = workbook['Edicao Manual']
            worksheet['A1'] = 'ID'
            worksheet['B1'] = 'valor incrementado'
            worksheet['C1'] = 'Data Modificação'
            worksheet['D1'] = 'Motivo'
            workbook.save(nomeplanilhaedmanual)
            
        workbook = load_workbook(nomeplanilhaedmanual)
        worksheet = workbook['Edicao Manual']
            
        for row in worksheet.iter_rows(min_col=1, max_col=4, min_row=1, max_row=1): # Colunas com dados
            #print('Linha: ',row)
            for cell in row: # Células com dados
                #print('Célula: ',cell)
                cell.font = Font(bold=True)  
                cell.border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
            
        workbook.save(nomeplanilhaedmanual)                     
        
        worksheet = workbook['Edicao Manual']
        worksheet['E1'] = ''
        worksheet['F1'] = '* Não apagar registros existentes'
        ajusta_largura(worksheet)
        workbook.save(nomeplanilhaedmanual)    
            
        workbook = protege_planilha(nomeplanilhaedmanual,'cnseg246$') # SENHA PLANILHA
        worksheet = workbook['Referencia MXM']
        workbook.save(nomeplanilhaedmanual)
        
        envia_arquivo(nomeplanilhaedmanual)
        s.close()
        
    except FileNotFoundError:
        print(f'O arquivo v_dre_real_qlik_{str(ano)}.csv não se encontra no diretório. O script Envia csv.py deve ser executado')
        s.close()
    
    except NaoAtualizaBD:
        print('A planilha de edição manual está sem registros, apesar do BD possuir. Os registros foram removidos indevidamente ? Utilizar os dados do BD ? (tbl_tmp_fluxo_caixa_edmanual ou tbl_tmp_dre_edmanual)')


def le_csv_dre(planilha,ano,mes): # DADOS DO ÚLTIMO MÊS COMPLETO
    
    df = read_csv(planilha, sep="|",dtype=object) # PARA EVITAR QUE O EXCEL CORTE OS LEADING ZEROS
    
    anomes = f'{str(ano)}{str(mes).zfill(2)}'
        
    df = df.loc[df['ANOMES'] == anomes].reset_index(drop=True) # OBTENDO OS REGISTROS DO MÊS QUE VÃO PASSAR POR MODIFICAÇÃO. MÊS ANTERIOR. 
    
    #print('Dataframe')
    #print(df)
    
    return df


def le_aba_edmanual(planilha):    
    
    df = read_excel(planilha,sheet_name='Edicao Manual',dtype=object) # PARA EVITAR QUE O EXCEL CORTE OS LEADING ZEROS
    
    df['Data Modificação'] = to_datetime(df['Data Modificação'],dayfirst=True).dt.strftime('%d/%m/%Y')    
    
    return df


def cria_planilha_fluxo_caixa(ano,mes): # A IDENTIFICAÇÃO DO REGISTRO MODIFICADO SERÁ PELA CONTA E PELO ANOMES
    
    global s

    s = conexao_sftp()                                
    s.chdir('./Projetos/Fluxo Caixa/Dados_Auxiliares/Planilha EdManual')   # temporarily chdir to bifiles
    
    if mes == None and ano == None:
       ano = date.today().year
       mes = date.today().month 
        
       if mes == 1:
        backup_planedmanuais(s,ano)
        
        ano = ano - 1
        mes = 12
        
       else:
        mes = mes - 1 # OS DADOS SÃO DO MÊS ANTERIOR.
    
    elif mes is not None and ano is not None:
        mes = mes
        ano = ano
                
       #else:
        #mes = mes - 1
    
    else:
        raise ParamError       
    
    print('Criando planilha de edição manual de Fluxo de Caixa...')
             
    nomeplanilhaedmanual = f'fluxo_caixa_edmanual_{ano}{str(mes).zfill(2)}.xlsx'
    
    try:
        df = le_csv_flxcx(f'v_fluxo_caixa_{ano}.csv',mes,ano) # PASSANDO O ANO E O MES ANTERIOR.
        
        # SE A PLANLHA EXISTIR, CRIA NOVAMENTE, SÓ QUE COM OS DADOS DA EDIÇÃO MANUAL.
        #files = os.listdir('./')
        files = s.listdir('./')        
        print('Arquivos no diretório de planilhas de edição manual: ',files)
            
        if nomeplanilhaedmanual in files:
            print('Planilha de edicao manual existente: ',nomeplanilhaedmanual)
            s.get(f'./{nomeplanilhaedmanual}')
            dfedmanual = le_aba_edmanual(nomeplanilhaedmanual)
            
            print('Aba Ed Manual')
            print(dfedmanual)
            
            # CRIA NOVAMENTE, SÓ QUE COM OS DADOS DA EDIÇÃO MANUAL.
            cria_planilha(df,nomeplanilhaedmanual) # CRIA AS ABAS REFERÊNCIA MXM E EDIÇÃO MANUAL
            
            with ExcelWriter(nomeplanilhaedmanual, engine='openpyxl',if_sheet_exists='replace',mode='a') as writer:
                dfedmanual.to_excel(writer, index=False, sheet_name='Edicao Manual') # INSERINDO DADOS DA EDIÇÃO MANUAL
                                
            inserir_no_bd('tbl_tmp_fluxo_caixa_edmanual',dfedmanual, 'fluxo_caixa') # PARA EFEITO DE TESTE, DEVE-SE ANTES FAZER TRUNCATE NESTA TABELA
                    
        else:
            workbook = cria_planilha(df,nomeplanilhaedmanual) # CRIA AS ABAS REFERÊNCIA MXM E EDIÇÃO MANUAL
            
            worksheet = workbook['Edicao Manual']
            worksheet['A1'] = 'CONTA'
            worksheet['B1'] = 'valor incrementado'
            worksheet['C1'] = 'Data Modificação'
            worksheet['D1'] = 'Motivo'
            workbook.save(nomeplanilhaedmanual)
            
                
        workbook = load_workbook(nomeplanilhaedmanual)
        worksheet = workbook['Edicao Manual']
            
        for row in worksheet.iter_rows(min_col=1, max_col=4, min_row=1, max_row=1): # Colunas com dados
            #print('Linha: ',row)
            for cell in row: # Células com dados
                #print('Célula: ',cell)
                cell.font = Font(bold=True)  
                cell.border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
            
        workbook.save(nomeplanilhaedmanual)                     
        
        worksheet = workbook['Edicao Manual']
        worksheet['E1'] = ''
        worksheet['F1'] = '* Não apagar registros existentes'
        ajusta_largura(worksheet)
        workbook.save(nomeplanilhaedmanual)    
            
        workbook = protege_planilha(nomeplanilhaedmanual,'cnseg246$') # SENHA PLANILHA
        worksheet = workbook['Referencia MXM']
        worksheet['E1'] = '* O sinal negativo foi adicionado, para refletir os valores das despesas no DFC'
        ajusta_largura(worksheet)
        
        workbook.save(nomeplanilhaedmanual)
        
        envia_arquivo(nomeplanilhaedmanual)    
        s.close()
        
    
    except FileNotFoundError:
        print(f'O arquivo v_fluxo_caixa_{ano}.csv não se encontra no diretório. O script Envia csv.py deve ser executado')
        s.close()
        
    except NaoAtualizaBD:
        print('A planilha de edição manual está sem registros, apesar do BD possuir. Os registros foram removidos indevidamente ? Utilizar os dados do BD ? (tbl_tmp_fluxo_caixa_edmanual ou tbl_tmp_dre_edmanual)')


def le_csv_flxcx(planilha,mes,ano): 
    
    
    df = read_csv(planilha, sep="|",dtype=object) # PARA EVITAR QUE O EXCEL CORTE OS LEADING ZEROS
    
    df = df[['CONTA','DESCRICAO',f'{str(mes).zfill(2)}/{ano}']] # OBTENDO OS REGISTROS DO MÊS QUE VÃO PASSAR POR MODIFICAÇÃO. MÊS ANTERIOR.
        
    for r in df[f'{str(mes).zfill(2)}/{ano}'].index:
        # COLOCANDO SINAL NEGATIVO, NO VALOR DAS CONTAS DE DESPESA, SE FOREM DIFERENTES DE ZERO
        if df.loc[r]['CONTA'][0:2] == '04' and str(df.loc[r,f'{str(mes).zfill(2)}/{ano}']).strip() != '0': # strip para remover espaços em branco
            df.loc[r,f'{str(mes).zfill(2)}/{ano}'] = '-'+str(df.loc[r,f'{str(mes).zfill(2)}/{ano}']).strip() # strip para remover espaços em branco
        else:
            df.loc[r,f'{str(mes).zfill(2)}/{ano}'] = str(df.loc[r,f'{str(mes).zfill(2)}/{ano}']).strip() # strip para remover espaços em branco
    
    return df


def protege_planilha(planilha,password):

    # Load the Excel workbook

    workbook = load_workbook(planilha)

    # Set the password to protect the workbook

    workbook.security.workbookPassword = password
    workbook.security.lockStructure = True
    
    worksheet = workbook['Referencia MXM']
    worksheet.protection.password = password
    worksheet.protection.sheet = True
    worksheet.protection.enable()
    
    worksheet = workbook['Edicao Manual']
    worksheet.protection.password = password
    worksheet.protection.sheet = True
    worksheet.protection.enable()
    
    # Save the workbook with the password applied

    workbook.save(planilha)
    
    return workbook

def backup_planedmanuais(s,ano):
    ano = ano - 2
    
    files = s.listdir('./')
    for f in files:
        if re.match(rf".*{ano}.*",f):
            print(f"Movendo arquivo {f} para backup")
            fromname = f
            toname = f"./backup/{fromname}"
            s.rename(fromname, toname) # MOVENDO O ARQUIVO



# BANCO PRD
senhabd = 'zvga2jd0871'
usuariobd = 'QLIK_MXM'
host = '10.10.20.61'
sid = 'MXMWEB2'

# ENVIO DOS ARQUIVOS
ip = '10.10.40.84'
username = 'qlik_admin'
password = '*wEd[<VTmb6uUql'


# A EXECUÇÃO DESSE SCRIPT SERÁ SOB DEMANDA, APÓS A RECEPÇÃO DO EMAIL DA ELIZANGELA INFORMANDO O FECHAMENTO CONTÁBIL.
def main():
    
    # SE QUISER FORÇAR UM ANO
    #ano = 2025
    ano = None
        
    # SE QUISER FORÇAR UM MÊS
    #mes = 3
    mes = None
    
    try:
        cria_planilha_dre(ano,mes)
        cria_planilha_fluxo_caixa(ano,mes)
    
    except(ParamError):
        print('Se forem passados parâmetros, o ano e o mês devem ser informados.') 
          
               
if __name__ == "__main__":        
    main()