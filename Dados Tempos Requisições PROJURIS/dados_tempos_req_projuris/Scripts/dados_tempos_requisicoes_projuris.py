'''
    A ORIGEM DOS DADOS É RELATÓRIO CUSTOMIZADO DE REQUISIÇÕES "Requisições Histórico - etapas".
'''	
from time import localtime
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border,Side
from pandas import ExcelWriter
from pandas import read_excel
from pandas import concat
from warnings import filterwarnings
from pandas.errors import SettingWithCopyWarning
import warnings

warnings.simplefilter(action='ignore', category=(SettingWithCopyWarning))
filterwarnings('ignore',category=UserWarning)

def ordenacao_porpeso(df):
    
    dictfase = {
                    'EM REQUISIÇÃO':1, # PESO
                    'AGUARDANDO EXECUÇÃO':2, # PESO
                    'EXECUTADA':3, # PESO
                    'FINALIZADA':4 # PESO
                }
    
    df['Fase da Requisição'] = df['Fase da Requisição'].map(dictfase)
    
    df.sort_values(by=['Número Requisição','Data','Fase da Requisição'],inplace=True)
    
    # RETORNANDO AOS VALORES ANTERIORES DE FASE
    df['Fase da Requisição'] = df['Fase da Requisição'].map({1:'EM REQUISIÇÃO',2:'AGUARDANDO EXECUÇÃO', 3:'EXECUTADA',4:'FINALIZADA'})
        
    
    return df

def filtros(df):
    
    # JÁ FILTRADO NO PROJURIS AS LINHAS DAS REQUISIÇÕES QUE POSSUEM AS FASES 'EM REQUISIÇÃO', 'AGUARDANDO EXECUÇÃO', 'EXECUTADA' e 'FINALIZADA'
    df = df.loc[df['Status Atual'] == 'FINALIZADA'].drop_duplicates()
    
    # REQUISIÇÕES ASSINADAS
    dfreqexecsassin = df.loc[(df['Fase da Requisição'] == 'EXECUTADA') & (df['Detalhes'].str.contains(r'[Aa]ssin',regex=True,case=False))]
    dfreqsassin = df.loc[df['Número Requisição'].isin(dfreqexecsassin['Número Requisição'])]    
    
    # REQUISIÇÕES ASSINADAS E QUE SÃO DE ADESÃO
    dfreqsadesao = df.loc[(df['Título'].str.contains(r'ADES[ÃA]O',regex=True,case=False)) | (df['Detalhes'].str.contains(r'ADES[ÃA]O',regex=True,case=False))]
    dfreqsassinadesao = dfreqsassin.loc[dfreqsassin['Número Requisição'].isin(dfreqsadesao['Número Requisição'])]
        
    # REQUISIÇÕES ASSINADAS E QUE NÃO SÃO DE ADESÃO
    df = df.loc[(df['Número Requisição'].isin(dfreqsassin['Número Requisição'])) & (~df['Número Requisição'].isin(dfreqsassinadesao['Número Requisição']))]    
        
    df = filtro_requisicoes_entrada_docusign(df)
    
    dfemrequisicao = filtro_momento_emrequisicao(df)    
    dfaguardaexec = filtro_momento_entrada_docusign(df)    
    dfexec = filtro_momento_execucao(df)
    dffinal = filtro_momento_final(df) 
            
    df = concat([dfemrequisicao,dfaguardaexec,dfexec,dffinal])
    
    return df


def filtro_requisicoes_entrada_docusign(df):
     
    dfreqsentradadocusign = df.loc[(df['Fase da Requisição'] == 'AGUARDANDO EXECUÇÃO') & (df['Detalhes'].str.contains(r'[Aa]ssin|[In]serido|[Dd]ocusign',regex=True,case=False))]
    df = df.loc[df['Número Requisição'].isin(dfreqsentradadocusign['Número Requisição'])]    

    return df


def filtro_momento_emrequisicao(df):

    dfemrequisicao = df.loc[df['Fase da Requisição'] == 'EM REQUISIÇÃO']
    min_indices_emrequisicao = dfemrequisicao.groupby('Número Requisição').apply(lambda x: x.index.min(),include_groups=False)
        
    dictreqsemrequisicao = dict(min_indices_emrequisicao)
    
    dfemrequisicao = dfemrequisicao.loc[dfemrequisicao.index.isin(dictreqsemrequisicao.values())]
    
    return dfemrequisicao


# FILTRO PARA OBTER OS MOMENTOS DE ENTRADA NO DOCUSIGN
def filtro_momento_entrada_docusign(df):
   
   dfaguardaexec = df.loc[df['Fase da Requisição'] == 'AGUARDANDO EXECUÇÃO']
   
   # OBTENDO O ÍNDICE MÁXIMO DE CADA REQUISIÇÃO QUE ESTÁ NA FASE 'AGUARDANDO EXECUÇÃO'
   max_indices_aguardaexec = dfaguardaexec.groupby('Número Requisição').apply(lambda x: x.index.max(),include_groups=False)
   dict_max_indices_aguardaexec = dict(max_indices_aguardaexec)

   dfaguardaexec = dfaguardaexec.loc[dfaguardaexec.index.isin(dict_max_indices_aguardaexec.values())]
    
   return dfaguardaexec


def filtro_momento_execucao(df):
    
    dfexec = df.loc[(df['Fase da Requisição'] == 'EXECUTADA') & (df['Detalhes'].str.contains(r'[Aa]ssin',regex=True,case=False))]
    
    max_indices_exec = dfexec.groupby('Número Requisição').apply(lambda x: x.index.max(),include_groups=False)
    dict_max_indices_exec = dict(max_indices_exec)
    
    dfexec = dfexec.loc[dfexec.index.isin(dict_max_indices_exec.values())]
    
    return dfexec 


def filtro_momento_final(df):
    
    dffinal = df.loc[df['Fase da Requisição'] == 'FINALIZADA']
    
    max_indices_exec = dffinal.groupby('Número Requisição').apply(lambda x: x.index.max(),include_groups=False)
    dict_max_indices_exec = dict(max_indices_exec)
    
    dffinal = dffinal.loc[dffinal.index.isin(dict_max_indices_exec.values())]
    
    return dffinal 


def calcula_tempos(df):
    
    dfdatasminreq = df.groupby('Número Requisição')[['Data']].min().rename(columns={'Data':'Data EM REQUISIÇÃO'}) 
    dfdatasmaxreq = df.groupby('Número Requisição')[['Data']].max().rename(columns={'Data':'Data EXECUTADA'})    
    
    dfdataentradadocusign = df.loc[df['Fase da Requisição'] == 'AGUARDANDO EXECUÇÃO'].rename(columns={'Data':'Data AGUARDANDO EXECUÇÃO'})     
      
    dfmerged = df[['Número Requisição','Requisitante','Tipo de Requisição']].set_index('Número Requisição').join(
                                                                                                                    [
                                                                                                                        dfdatasminreq,
                                                                                                                        dfdataentradadocusign[['Número Requisição','Data AGUARDANDO EXECUÇÃO']].set_index('Número Requisição'),
                                                                                                                        dfdatasmaxreq
                                                                                                                    ]
                                                                                                                ).drop_duplicates()
    
    dfmerged.insert(loc=5,column='Tempo Assinatura',value=(dfmerged['Data EXECUTADA'] - dfmerged['Data AGUARDANDO EXECUÇÃO']))
        
    # FAZENDO CÁLCULO COM DATAFRAMES    
    dfmerged['Tempo Total em dias'] = dfmerged['Data EXECUTADA'] - dfmerged['Data EM REQUISIÇÃO']   
    
    return dfmerged
        
def cria_planilha(dfold,df,dftempos):    
    
    data = f'{str(localtime().tm_mday).zfill(2)}{str(localtime().tm_mon).zfill(2)}{localtime().tm_year}{str(localtime().tm_hour).zfill(2)}{str(localtime().tm_min).zfill(2)}{str(localtime().tm_sec).zfill(2)}'
        
    outputfilename = f'Requisições Histórico - etapas_saida {data}.xlsx'
    
    with ExcelWriter(dirpath+outputfilename, engine='openpyxl') as writer:
        dfold.to_excel(writer,index=False,sheet_name='Relatório Projuris Web')
        df.to_excel(writer,index=False,sheet_name='Dados limpos')    
        dftempos.to_excel(writer,sheet_name='Tempos')
    
    wb = load_workbook(dirpath+outputfilename)
    
    ws = wb['Relatório Projuris Web']    
    ajusta_largura(ws)
    ws.freeze_panes = 'A2'  
            
    ws = wb['Dados limpos']    
    ajusta_largura(ws)
    ws.freeze_panes = 'A2'   
    
    ws = wb['Tempos']
    ajusta_largura(ws)
    ws.freeze_panes = 'A2' # CONGELAMENTO DA LINHA SUPERIOR    
    # RETIRANDO A BORDA
    for cell in ws['A2':(ws.cell(row=ws.max_row,column=ws.min_column)).coordinate]: 
        cell[0].border = Border(bottom=Side(style='none'),right=Side(style='none'),left=Side(style='none'),top=Side(style='none'))
        
    wb.active = wb['Tempos']
    
    wb.save(dirpath+outputfilename)
    
    return wb
    
def ajusta_largura(worksheet):
    
    for col in worksheet.columns:
        max_length = 0        
        for cell in col: 
            #print('Coluna: ',cell.column_letter)
            if isinstance(cell.value, str):
                max_length = max(max_length, len(str(cell.value))) # RETORNA O MAIOR TAMANHO ENTRE 2 TAMANHOS
            elif isinstance(cell.value, int) or isinstance(cell.value, float):
                max_length = max(max_length, len(str(cell.value)))            
                
            # QUEBRA DE LINHA. SÓ CONSEGUI FAZENDO PARA TODAS AS CÉLULAS DA PLANILHA
            cell.alignment = Alignment(wrapText=True,vertical='top',horizontal='left')        
        
        #print('1 col da tupla: ',col)
        col = col[0].column_letter
                
        if col == 'F' and worksheet.title in ['Dados limpos','Relatório Projuris Web']: # Data
            worksheet.column_dimensions[col].width = max_length + 16            
        elif col in ['B','H'] and (worksheet.title in ['Dados limpos','Relatório Projuris Web']): # Título e Detalhes
            worksheet.column_dimensions[col].width = 100
            
        elif col in ['D','E','F'] and worksheet.title == 'Tempos': # Data EM REQUISIÇÃO, Data AGUARDANDO EXECUÇÃO, Data EXECUTADA
            worksheet.column_dimensions[col].width = max_length + 9
        
        else:
            worksheet.column_dimensions[col].width = max_length + 2    
    
    return worksheet


#dirpath = 'C:\\Users\\anton\\OneDrive - FEDERACAO NACIONAL DAS EMP DE SEG PRIVADOS E DE CAPITALIZACAO\\Documentos\\Sistema\\Projuris\\Relatório Tempos em Requisições\\'
dirpath = '.\\' # cxfreeze dados_tempos_requisicoes_projuris.py --target-dir ..\executavel. O executável será chamado por um atalho
relatorioprojuris = 'Requisições Histórico - etapas.xlsx'

def main():
    
    print('Executando. Aguarde...')
    
    dfold = read_excel(dirpath+relatorioprojuris)
    
    dfold.rename(columns={'Status':'Status Atual'},inplace=True)
    
    df = dfold.drop_duplicates(subset=['Número Requisição','Detalhes'])
    df = filtros(df)
    df = ordenacao_porpeso(df)
    
    dftempos = calcula_tempos(df)
        
    cria_planilha(dfold,df,dftempos)       
        
    
if __name__ == "__main__":        
    main()